"""
Pagina: Modelo financiero de inversion en alerta temprana (EWS).

Analisis costo-beneficio de sistemas de alerta temprana frente a
perdidas historicas por catastrofe, con escenarios, proyecciones a
10 anios y analisis critico de sesgos metodologicos.
"""

import base64
import sys
from pathlib import Path

import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

sys.path.append(str(Path(__file__).parent.parent))
from theme import inyectar_tema, encabezado_pagina, pie_sidebar, CHART_PALETTE, CHART_SCALE_RIESGO, CHART_SCALE_TEAL

ASSETS = Path(__file__).parent.parent / "assets"
XLSX = ASSETS / "Modelo_Catastrofes_Alerta_Temprana.xlsx"
PDF_RESUMEN = ASSETS / "Resumen_Ejecutivo_Catastrofes_EWS.pdf"
PDF_SESGOS = ASSETS / "Analisis_Sesgos_Modelo_Catastrofes.pdf"
IMG_INFOGRAFIA = ASSETS / "infografia_completa.png"
IMG_SESGOS = ASSETS / "infografia_sesgos.png"
IMG_CHART = ASSETS / "chart_escenarios.png"

PLOT_BG = dict(plot_bgcolor="#F7F6F2", paper_bgcolor="#F7F6F2")

inyectar_tema()


# ── Carga de datos desde Excel ───────────────────────────

@st.cache_data
def get_resumen(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Resumen"]
    rows = []
    for r in range(8, 17):
        region = ws.cell(r, 2).value
        if region is None:
            continue
        rows.append({
            "Region": region,
            "Perdidas Hist. ($B)": ws.cell(r, 3).value or 0,
            "Inversion EWS ($B)": ws.cell(r, 4).value or 0,
            "Perdidas Evitadas 10a ($B)": ws.cell(r, 5).value or 0,
            "Ahorro FR ($B)": ws.cell(r, 6).value or 0,
            "Beneficio Total ($B)": ws.cell(r, 7).value or 0,
            "Coste Total EWS ($B)": ws.cell(r, 8).value or 0,
            "ROI": ws.cell(r, 9).value or 0,
            "BCR": ws.cell(r, 10).value or 0,
            "Reduccion % PIB": ws.cell(r, 11).value or 0,
        })
    return pd.DataFrame(rows)


@st.cache_data
def get_escenarios(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Escenarios"]
    data = {}
    for r in range(15, 24):
        metric = ws.cell(r, 2).value
        if metric is None:
            continue
        data[metric] = {
            "Conservador": ws.cell(r, 3).value, "Base": ws.cell(r, 4).value,
            "Optimista": ws.cell(r, 5).value, "Unidad": ws.cell(r, 6).value,
        }
    return data


@st.cache_data
def get_historicos(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Datos Históricos"]
    rows = []
    for r in range(6, 14):
        region = ws.cell(r, 2).value
        if region is None:
            continue
        rows.append({
            "Region": region, "2021": ws.cell(r, 3).value or 0, "2022": ws.cell(r, 4).value or 0,
            "2023": ws.cell(r, 5).value or 0, "2024": ws.cell(r, 6).value or 0,
            "2025": ws.cell(r, 7).value or 0, "Promedio": ws.cell(r, 9).value or 0,
        })
    return pd.DataFrame(rows)


@st.cache_data
def get_proyecciones(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Proyecciones"]
    years = list(range(2026, 2036))
    regiones = ["Norteamerica", "Sudamerica", "Asia", "Europa Occidental",
                "Europa Meridional", "Europa Oriental", "Africa", "Oceania y Resto"]
    filas = []
    for i, region in enumerate(regiones):
        net_row = 20 + i * 6 + 4
        fila = {"Region": region}
        for j, year in enumerate(years):
            val = ws.cell(net_row, 3 + j).value
            fila[str(year)] = val if val is not None else 0
        filas.append(fila)
    return pd.DataFrame(filas)


def file_download(path, label):
    if not Path(path).exists():
        st.sidebar.caption(f"No disponible: {Path(path).name}")
        return
    st.download_button(label, Path(path).read_bytes(), file_name=Path(path).name, mime="application/octet-stream")


def display_pdf(path):
    if not Path(path).exists():
        st.info(f"Documento no disponible: {Path(path).name}")
        return
    with open(path, "rb") as f:
        pdf_bytes = base64.b64encode(f.read()).decode()
    st.markdown(
        f'<iframe src="data:application/pdf;base64,{pdf_bytes}" width="100%" height="880px" type="application/pdf"></iframe>',
        unsafe_allow_html=True,
    )


def mostrar_imagen(path, caption):
    if Path(path).exists():
        st.image(str(path), caption=caption, use_container_width=True)
    else:
        st.info(f"Imagen no disponible: {Path(path).name}")


df_resumen = get_resumen(XLSX)
df_historicos = get_historicos(XLSX)
escenarios = get_escenarios(XLSX)
df_proyecciones = get_proyecciones(XLSX)

df_regiones = df_resumen[df_resumen["Region"] != "TOTAL GLOBAL"].copy()
total_row = df_resumen[df_resumen["Region"] == "TOTAL GLOBAL"].iloc[0]


# ── Sidebar ───────────────────────────────────────────────

st.sidebar.title("Modelo financiero")
st.sidebar.caption("Catastrofes naturales vs inversion en alerta temprana (EWS).")
st.sidebar.markdown("---")
st.sidebar.markdown("**Fuentes del modelo**")
st.sidebar.caption("Aon, Swiss Re, Munich Re, EM-DAT, WMO/UNDRR, World Bank")
st.sidebar.markdown("---")
st.sidebar.markdown("**Descargas**")
file_download(XLSX, "Descargar modelo (Excel)")
file_download(PDF_RESUMEN, "Resumen ejecutivo (PDF)")
file_download(PDF_SESGOS, "Analisis de sesgos (PDF)")
pie_sidebar("Modelo financiero EWS")


# ── Layout ────────────────────────────────────────────────

encabezado_pagina(
    "Analisis de inversion",
    "Catastrofes naturales frente a inversion en alerta temprana",
    "Modelo cuantitativo del retorno de invertir en sistemas de alerta temprana (EWS) por region, "
    "con datos historicos, escenarios de sensibilidad y proyecciones a 10 anios.",
)

tab_resumen, tab_financiero, tab_historicas, tab_proyecciones, tab_escenarios, tab_docs, tab_sesgos = st.tabs([
    "Resumen ejecutivo", "Modelo financiero", "Perdidas historicas",
    "Proyecciones a 10 anios", "Comparativa de escenarios", "Documentacion", "Analisis de sesgos",
])

with tab_resumen:
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Perdidas anuales globales", f"${total_row['Perdidas Hist. ($B)']:.0f}B", "Promedio 2021-2025")
    col2.metric("Inversion EWS propuesta", f"${total_row['Inversion EWS ($B)']:.1f}B", "Global, 8 regiones")
    col3.metric("Ratio beneficio-coste", f"{total_row['BCR']:.1f}x", "Cada $1 invertido")
    col4.metric("ROI proyectado", f"{total_row['ROI']*100:.0f}%", "A 10 anios")

    col5, col6, col7, col8 = st.columns(4)
    col5.metric("Perdidas evitadas (10 anios)", f"${total_row['Perdidas Evitadas 10a ($B)']:.0f}B")
    col6.metric("Ahorro en primera respuesta", f"${total_row['Ahorro FR ($B)']:.1f}B")
    col7.metric("Beneficio total (10 anios)", f"${total_row['Beneficio Total ($B)']:.0f}B")
    col8.metric("Reduccion del PIB global", f"{total_row['Reduccion % PIB']*100:.2f}%")

    st.markdown("---")
    st.markdown(f"""
    <div class="callout">
    <h4>Hallazgo principal</h4>
    <p>Una inversion global de <b>${total_row['Inversion EWS ($B)']:.1f} mil millones</b> en sistemas de
    alerta temprana genera un beneficio estimado de <b>${total_row['Beneficio Total ($B)']:.0f} mil millones</b>
    en 10 anios &mdash; un ratio beneficio-coste de <b>{total_row['BCR']:.1f}x</b>, consistente con el rango
    de 3:1 a 10:1 documentado por McKinsey y Naciones Unidas.</p>
    </div>
    """, unsafe_allow_html=True)

    st.subheader("Ratio beneficio-coste por region")
    df_sorted = df_regiones.sort_values("BCR", ascending=True)
    fig = px.bar(df_sorted, x="BCR", y="Region", orientation="h", color="BCR",
                 color_continuous_scale=CHART_SCALE_RIESGO[::-1],
                 text=df_sorted["BCR"].apply(lambda x: f"{x:.1f}x"))
    fig.update_layout(height=380, showlegend=False, **PLOT_BG)
    fig.update_traces(textposition="outside")
    fig.update_coloraxes(showscale=False)
    st.plotly_chart(fig, use_container_width=True)

    st.subheader("Infografia de resumen")
    mostrar_imagen(IMG_INFOGRAFIA, "Sintesis visual del modelo de inversion en alerta temprana")

with tab_financiero:
    st.subheader("Indicadores por region")
    df_display = df_resumen.copy()
    for col in ["Perdidas Hist. ($B)", "Inversion EWS ($B)", "Perdidas Evitadas 10a ($B)",
                "Ahorro FR ($B)", "Beneficio Total ($B)", "Coste Total EWS ($B)"]:
        df_display[col] = df_display[col].apply(lambda x: f"${x:.1f}B")
    df_display["ROI"] = df_display["ROI"].apply(lambda x: f"{x*100:.0f}%")
    df_display["BCR"] = df_display["BCR"].apply(lambda x: f"{x:.1f}x")
    df_display["Reduccion % PIB"] = df_display["Reduccion % PIB"].apply(lambda x: f"{x*100:.2f}%")
    st.dataframe(df_display, use_container_width=True, hide_index=True)

    st.markdown("---")
    col_a, col_b = st.columns(2)
    with col_a:
        st.subheader("Inversion vs beneficio total")
        fig2 = go.Figure()
        fig2.add_trace(go.Bar(name="Inversion EWS", x=df_regiones["Region"], y=df_regiones["Inversion EWS ($B)"], marker_color=CHART_PALETTE[1]))
        fig2.add_trace(go.Bar(name="Beneficio total", x=df_regiones["Region"], y=df_regiones["Beneficio Total ($B)"], marker_color=CHART_PALETTE[0]))
        fig2.update_layout(barmode="group", height=380, **PLOT_BG)
        st.plotly_chart(fig2, use_container_width=True)
    with col_b:
        st.subheader("ROI por region")
        df_roi = df_regiones.sort_values("ROI", ascending=False)
        fig3 = px.bar(df_roi, x="Region", y="ROI", color="ROI", color_continuous_scale=CHART_SCALE_RIESGO[::-1],
                      text=df_roi["ROI"].apply(lambda x: f"{x*100:.0f}%"))
        fig3.update_layout(height=380, showlegend=False, **PLOT_BG)
        fig3.update_traces(textposition="outside")
        fig3.update_coloraxes(showscale=False)
        st.plotly_chart(fig3, use_container_width=True)

    st.subheader("Perdidas evitadas frente al coste de EWS")
    fig4 = go.Figure()
    fig4.add_trace(go.Bar(name="Perdidas evitadas (10 anios)", x=df_regiones["Region"], y=df_regiones["Perdidas Evitadas 10a ($B)"], marker_color=CHART_PALETTE[0]))
    fig4.add_trace(go.Bar(name="Ahorro primera respuesta", x=df_regiones["Region"], y=df_regiones["Ahorro FR ($B)"], marker_color=CHART_PALETTE[3]))
    fig4.add_trace(go.Bar(name="Coste total EWS", x=df_regiones["Region"], y=df_regiones["Coste Total EWS ($B)"], marker_color=CHART_PALETTE[1]))
    fig4.update_layout(barmode="group", height=420, **PLOT_BG)
    st.plotly_chart(fig4, use_container_width=True)

with tab_historicas:
    st.subheader("Perdidas historicas por region, 2021-2025 ($B)")
    st.dataframe(df_historicos, use_container_width=True, hide_index=True)

    df_melt = df_historicos.melt(id_vars=["Region"], value_vars=["2021", "2022", "2023", "2024", "2025"],
                                  var_name="Anio", value_name="Perdidas ($B)")
    st.subheader("Evolucion de perdidas por region")
    fig5 = px.line(df_melt, x="Anio", y="Perdidas ($B)", color="Region", markers=True, color_discrete_sequence=CHART_PALETTE)
    fig5.update_layout(height=420, **PLOT_BG)
    st.plotly_chart(fig5, use_container_width=True)

    st.subheader("Composicion de perdidas por anio")
    fig6 = px.bar(df_melt, x="Anio", y="Perdidas ($B)", color="Region", color_discrete_sequence=CHART_PALETTE)
    fig6.update_layout(height=380, **PLOT_BG)
    st.plotly_chart(fig6, use_container_width=True)

    st.subheader("Promedio de perdidas por region")
    df_prom = df_historicos.sort_values("Promedio", ascending=True)
    fig7 = px.bar(df_prom, x="Promedio", y="Region", orientation="h", color="Promedio",
                  color_continuous_scale=CHART_SCALE_TEAL, text=df_prom["Promedio"].apply(lambda x: f"${x:.1f}B"))
    fig7.update_layout(height=380, showlegend=False, **PLOT_BG)
    fig7.update_traces(textposition="outside")
    fig7.update_coloraxes(showscale=False)
    st.plotly_chart(fig7, use_container_width=True)

with tab_proyecciones:
    st.subheader("Flujo neto anual, 2026-2035 (escenario base)")
    region_sel = st.selectbox("Region", df_proyecciones["Region"].tolist())
    row_sel = df_proyecciones[df_proyecciones["Region"] == region_sel].iloc[0]
    years = [str(y) for y in range(2026, 2036)]
    valores = [row_sel[y] for y in years]

    col1, col2, col3 = st.columns(3)
    col1.metric("Flujo neto total (10 anios)", f"${sum(valores):.1f}B")
    col2.metric("Flujo neto promedio anual", f"${sum(valores)/10:.1f}B")
    col3.metric("Mejor anio", f"${max(valores):.1f}B", str(2026 + valores.index(max(valores))))

    fig8 = go.Figure()
    fig8.add_trace(go.Bar(x=years, y=valores, marker_color="#01696F",
                          text=[f"${v:.1f}" for v in valores], textposition="outside"))
    fig8.update_layout(title=f"Flujo neto anual — {region_sel}", xaxis_title="Anio", yaxis_title="$B",
                       height=380, **PLOT_BG)
    st.plotly_chart(fig8, use_container_width=True)

    st.subheader("Flujo neto por region y anio")
    df_heat = df_proyecciones.set_index("Region")
    fig9 = px.imshow(df_heat, color_continuous_scale=CHART_SCALE_RIESGO[::-1],
                     labels=dict(x="Anio", y="Region", color="$B"), text_auto=".1f")
    fig9.update_layout(height=420, **PLOT_BG)
    st.plotly_chart(fig9, use_container_width=True)

    st.subheader("Tabla detallada de proyecciones")
    st.dataframe(df_proyecciones, use_container_width=True, hide_index=True)

with tab_escenarios:
    st.subheader("Conservador, base y optimista")
    esc_data = [{"Metrica": m, **v} for m, v in escenarios.items()]
    st.dataframe(pd.DataFrame(esc_data), use_container_width=True, hide_index=True)

    col_a, col_b = st.columns(2)
    with col_a:
        st.subheader("Ratio beneficio-coste por escenario")
        bcr = escenarios["BCR (Beneficio/Coste)"]
        fig10 = go.Figure()
        fig10.add_trace(go.Bar(x=["Conservador", "Base", "Optimista"],
                               y=[bcr["Conservador"], bcr["Base"], bcr["Optimista"]],
                               marker_color=["#A84B2F", "#01696F", "#3A6B1E"],
                               text=[f'{bcr[k]:.1f}x' for k in ["Conservador", "Base", "Optimista"]],
                               textposition="outside"))
        fig10.update_layout(height=380, yaxis_title="BCR (x)", **PLOT_BG)
        fig10.add_hline(y=3, line_dash="dash", line_color="#7A3714", annotation_text="Umbral de referencia (3:1)")
        st.plotly_chart(fig10, use_container_width=True)
    with col_b:
        st.subheader("ROI por escenario")
        roi = escenarios["ROI proyectado"]
        fig11 = go.Figure()
        fig11.add_trace(go.Bar(x=["Conservador", "Base", "Optimista"],
                               y=[roi[k]*100 for k in ["Conservador", "Base", "Optimista"]],
                               marker_color=["#A84B2F", "#01696F", "#3A6B1E"],
                               text=[f'{roi[k]*100:.0f}%' for k in ["Conservador", "Base", "Optimista"]],
                               textposition="outside"))
        fig11.update_layout(height=380, yaxis_title="ROI (%)", **PLOT_BG)
        st.plotly_chart(fig11, use_container_width=True)

    st.subheader("Beneficio frente a coste por escenario")
    ben, cost = escenarios["Beneficio total 10 años ($B)"], escenarios["Coste total EWS ($B)"]
    fig12 = go.Figure()
    fig12.add_trace(go.Bar(name="Beneficio total", x=["Conservador", "Base", "Optimista"],
                           y=[ben[k] for k in ["Conservador", "Base", "Optimista"]], marker_color="#01696F"))
    fig12.add_trace(go.Bar(name="Coste total EWS", x=["Conservador", "Base", "Optimista"],
                           y=[cost[k] for k in ["Conservador", "Base", "Optimista"]], marker_color="#A84B2F"))
    fig12.update_layout(barmode="group", height=380, yaxis_title="$B", **PLOT_BG)
    st.plotly_chart(fig12, use_container_width=True)

    st.subheader("Grafico comparativo de referencia")
    mostrar_imagen(IMG_CHART, "Comparativa de escenarios en cuatro paneles")

with tab_docs:
    st.subheader("Resumen ejecutivo")
    st.caption("Documento de cuatro paginas con indicadores, metodologia, escenarios y conclusiones.")
    display_pdf(PDF_RESUMEN)
    st.markdown("---")
    st.subheader("Modelo financiero completo")
    st.caption("Seis hojas: Resumen, Supuestos, Datos historicos, Escenarios, Proyecciones y Metodologia.")
    file_download(XLSX, "Descargar Excel completo")

with tab_sesgos:
    st.markdown("""
    <div class="callout-warn">
    <h4>Nota metodologica</h4>
    <p>Un modelo que reconoce sus limitaciones es mas creible que uno que las oculta. El ratio
    beneficio-coste conservador de 3.4x se mantiene por encima del umbral de referencia de 3:1
    incluso tras este analisis critico.</p>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Desastres sin datos de danios", "41.5%", "EM-DAT")
    col2.metric("Sin danios asegurados registrados", "88.1%", "EM-DAT")
    col3.metric("Reduccion de mortalidad con EWS", "6 a 8x", "No incluido en el ROI")
    col4.metric("PIB perdido, Jamaica", "40%", "Un solo huracan")

    st.subheader("Infografia de sesgos")
    mostrar_imagen(IMG_SESGOS, "Sintesis visual de los seis sesgos criticos del modelo")

    st.subheader("Analisis completo de sesgos")
    st.caption("Siete paginas con analisis detallado de cada sesgo, fuentes academicas y mitigaciones.")
    display_pdf(PDF_SESGOS)

    st.markdown("""
    <div class="callout">
    <h4>Conclusion</h4>
    <p>Reconocer los sesgos metodologicos no debilita el caso de inversion: incluso en el escenario
    conservador, el ratio de 3.4x supera el umbral minimo de 3:1. Incluir estas salvedades fortalece
    la credibilidad del modelo y favorece una toma de decisiones mas informada.</p>
    </div>
    """, unsafe_allow_html=True)