"""
Dashboard Streamlit — Catástrofes Naturales vs Inversión en Alerta Temprana
Modelo financiero, escenarios, infografías y análisis de sesgos.

Ejecutar:  streamlit run app.py
"""

import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import pandas as pd
import openpyxl
import base64
from pathlib import Path

# ──────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────
st.set_page_config(
    page_title="Catástrofes vs Alerta Temprana",
    page_icon="🌍",
    layout="wide",
    initial_sidebar_state="expanded",
)

ASSETS = Path(__file__).parent / "assets"
XLSX = ASSETS / "Modelo_Catastrofes_Alerta_Temprana.xlsx"
PDF_RESUMEN = ASSETS / "Resumen_Ejecutivo_Catastrofes_EWS.pdf"
PDF_SESGOS = ASSETS / "Analisis_Sesgos_Modelo_Catastrofes.pdf"
IMG_INFOGRAFIA = ASSETS / "infografia_completa.png"
IMG_SESGOS = ASSETS / "infografia_sesgos.png"
IMG_CHART = ASSETS / "chart_escenarios.png"

# Colores (paleta Nexus teal)
COLORS = {
    "bg": "#F7F6F2",
    "surface": "#F9F8F5",
    "text": "#28251D",
    "muted": "#7A7974",
    "primary": "#01696F",
    "primary_dark": "#0C4E54",
    "accent": "#20808D",
    "terra": "#A84B2F",
    "gold": "#D19900",
    "mauve": "#944454",
    "success": "#437A22",
    "warning": "#964219",
}
CHART_PALETTE = ["#20808D", "#A84B2F", "#1B474D", "#BCE2E7", "#944454", "#FFC553", "#848456", "#6E522B"]

# ──────────────────────────────────────────────
# CSS
# ──────────────────────────────────────────────
st.markdown("""
<style>
    /* Fondo general */
    .stApp { background-color: #F7F6F2; }

    /* Headers */
    h1, h2, h3 { color: #01696F !important; font-family: 'DM Sans', 'Inter', sans-serif; }

    /* KPI cards */
    div[data-testid="stMetric"] {
        background: #F9F8F5;
        border: 1px solid #D4D1CA;
        border-radius: 10px;
        padding: 16px 20px;
    }
    div[data-testid="stMetric"] label {
        color: #7A7974;
        font-size: 0.85rem;
    }
    div[data-testid="stMetric"] div[data-testid="stMetricValue"] {
        color: #01696F;
        font-weight: 800;
    }

    /* Sidebar */
    section[data-testid="stSidebar"] {
        background: #0D1B2A;
    }
    section[data-testid="stSidebar"] * {
        color: #E0E1DD;
    }
    section[data-testid="stSidebar"] .stSelectbox label,
    section[data-testid="stSidebar"] .stRadio label {
        color: #9AB0B4;
    }

    /* Botones descarga */
    .stDownloadButton > button {
        background: #01696F;
        color: white;
        border: none;
        border-radius: 8px;
        padding: 8px 24px;
        font-weight: 600;
    }
    .stDownloadButton > button:hover {
        background: #0C4E54;
    }

    /* Callout */
    .callout {
        background: #F9F8F5;
        border-left: 4px solid #01696F;
        border-radius: 8px;
        padding: 16px 20px;
        margin: 12px 0;
    }
    .callout-warn {
        background: #FFF8F0;
        border-left: 4px solid #964219;
        border-radius: 8px;
        padding: 16px 20px;
        margin: 12px 0;
    }

    /* Tabla */
    .dataframe { border-radius: 8px; overflow: hidden; }
    .dataframe th { background: #01696F !important; color: white !important; }
    .dataframe td { background: #F9F8F5 !important; }
</style>
""", unsafe_allow_html=True)


# ──────────────────────────────────────────────
# DATA LOADING
# ──────────────────────────────────────────────
@st.cache_data
def load_workbook(path):
    return openpyxl.load_workbook(path, data_only=True)

@st.cache_data
def get_resumen(path):
    """Lee la hoja Resumen del Excel con valores calculados."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Resumen"]
    rows = []
    for r in range(8, 17):  # filas 8-16 (regiones + total)
        region = ws.cell(r, 2).value
        if region is None:
            continue
        rows.append({
            "Región": region,
            "Pérdidas Hist. ($B)": ws.cell(r, 3).value or 0,
            "Inversión EWS ($B)": ws.cell(r, 4).value or 0,
            "Pérdidas Evitadas 10a ($B)": ws.cell(r, 5).value or 0,
            "Ahorro FR ($B)": ws.cell(r, 6).value or 0,
            "Beneficio Total ($B)": ws.cell(r, 7).value or 0,
            "Coste Total EWS ($B)": ws.cell(r, 8).value or 0,
            "ROI": ws.cell(r, 9).value or 0,
            "BCR": ws.cell(r, 10).value or 0,
            "Reducción % PIB": ws.cell(r, 11).value or 0,
        })
    return pd.DataFrame(rows)

@st.cache_data
def get_escenarios(path):
    """Lee la hoja Escenarios."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Escenarios"]
    data = {}
    for r in range(15, 24):
        metric = ws.cell(r, 2).value
        if metric is None:
            continue
        data[metric] = {
            "Conservador": ws.cell(r, 3).value,
            "Base": ws.cell(r, 4).value,
            "Optimista": ws.cell(r, 5).value,
            "Unidad": ws.cell(r, 6).value,
        }
    return data

@st.cache_data
def get_historicos(path):
    """Lee la hoja Datos Históricos."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Datos Históricos"]
    rows = []
    for r in range(6, 14):
        region = ws.cell(r, 2).value
        if region is None:
            continue
        rows.append({
            "Región": region,
            "2021": ws.cell(r, 3).value or 0,
            "2022": ws.cell(r, 4).value or 0,
            "2023": ws.cell(r, 5).value or 0,
            "2024": ws.cell(r, 6).value or 0,
            "2025": ws.cell(r, 7).value or 0,
            "Promedio": ws.cell(r, 9).value or 0,
        })
    return pd.DataFrame(rows)

@st.cache_data
def get_supuestos(path):
    """Lee la hoja Supuestos."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Supuestos"]
    rows = []
    for r in range(26, 35):
        region = ws.cell(r, 2).value
        if region is None:
            continue
        rows.append({
            "Región": region,
            "Inversión EWS ($B)": ws.cell(r, 3).value or 0,
            "O&M Anual ($B)": ws.cell(r, 4).value or 0,
            "PIB ($B)": ws.cell(r, 5).value or 0,
        })
    return pd.DataFrame(rows)

@st.cache_data
def get_proyecciones(path):
    """Lee la hoja Proyecciones (flujo neto por región y año)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Proyecciones"]
    years = list(range(2026, 2036))
    regions = ["Norteamérica", "Sudamérica", "Asia", "Europa Occidental",
               "Europa Meridional", "Europa Oriental", "África", "Oceanía y Resto"]
    all_rows = []
    for i, region in enumerate(regions):
        base_row = 20 + i * 6  # filas: 20, 26, 32, 38, 44, 50, 56, 62
        # Flujo neto está 4 filas debajo del header de región
        net_row = base_row + 4
        row_data = {"Región": region}
        for j, year in enumerate(years):
            col = 3 + j  # columnas C=3 hasta L=12
            val = ws.cell(net_row, col).value
            row_data[str(year)] = val if val is not None else 0
        all_rows.append(row_data)
    return pd.DataFrame(all_rows)


# ──────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────
def file_download(path, label):
    """Crea un botón de descarga para cualquier archivo."""
    data = Path(path).read_bytes()
    st.download_button(label, data, file_name=Path(path).name, mime="application/octet-stream")

def display_pdf(path):
    """Muestra un PDF embebido en base64."""
    with open(path, "rb") as f:
        pdf_bytes = base64.b64encode(f.read()).decode()
    st.markdown(f'<iframe src="data:application/pdf;base64,{pdf_bytes}" width="100%" height="900px" type="application/pdf"></iframe>', unsafe_allow_html=True)


# ──────────────────────────────────────────────
# LOAD DATA
# ──────────────────────────────────────────────
df_resumen = get_resumen(XLSX)
df_historicos = get_historicos(XLSX)
escenarios = get_escenarios(XLSX)
df_proyecciones = get_proyecciones(XLSX)

# Separar total de regiones
df_regiones = df_resumen[df_resumen["Región"] != "TOTAL GLOBAL"].copy()
total_row = df_resumen[df_resumen["Región"] == "TOTAL GLOBAL"].iloc[0]


# ──────────────────────────────────────────────
# SIDEBAR
# ──────────────────────────────────────────────
st.sidebar.title("🌍 Catástrofes vs EWS")
st.sidebar.markdown("---")
st.sidebar.markdown("**Modelo de inversión en Alerta Temprana**")
st.sidebar.markdown("Datos base: Aon, Swiss Re, Munich Re, EM-DAT, WMO/UNDRR")
st.sidebar.markdown("---")

pagina = st.sidebar.radio("Navegación", [
    "📊 Resumen Ejecutivo",
    "💰 Modelo Financiero",
    "📉 Pérdidas Históricas",
    "🔮 Proyecciones 10 años",
    "📈 Comparativa de Escenarios",
    "📄 Documentación",
    "⚠️ Análisis de Sesgos",
])

st.sidebar.markdown("---")
st.sidebar.markdown("**Descargas**")
file_download(XLSX, "📊 Descargar Excel")
file_download(PDF_RESUMEN, "📄 PDF Resumen")
file_download(PDF_SESGOS, "⚠️ PDF Sesgos")
file_download(IMG_INFOGRAFIA, "🖼️ Infografía")
file_download(IMG_SESGOS, "⚠️ Infografía Sesgos")

st.sidebar.markdown("---")
st.sidebar.caption("© 2026 | Perplexity Computer")


# ──────────────────────────────────────────────
# PÁGINA: RESUMEN EJECUTIVO
# ──────────────────────────────────────────────
if pagina == "📊 Resumen Ejecutivo":
    st.title("📊 Resumen Ejecutivo")
    st.markdown("### Catástrofes Naturales vs Inversión en Alerta Temprana (EWS)")
    st.markdown("---")

    # KPIs principales
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Pérdidas anuales globales", f"${total_row['Pérdidas Hist. ($B)']:.0f}B", "Promedio 2021-2025")
    with col2:
        st.metric("Inversión EWS propuesta", f"${total_row['Inversión EWS ($B)']:.1f}B", "Global, 8 regiones")
    with col3:
        st.metric("BCR (Beneficio/Coste)", f"{total_row['BCR']:.1f}x", "Cada $1 → $7.2")
    with col4:
        st.metric("ROI proyectado", f"{total_row['ROI']*100:.0f}%", "10 años")

    st.markdown("---")

    col5, col6, col7, col8 = st.columns(4)
    with col5:
        st.metric("Pérdidas evitadas (10 años)", f"${total_row['Pérdidas Evitadas 10a ($B)']:.0f}B")
    with col6:
        st.metric("Ahorro First Responders", f"${total_row['Ahorro FR ($B)']:.1f}B")
    with col7:
        st.metric("Beneficio total (10 años)", f"${total_row['Beneficio Total ($B)']:.0f}B")
    with col8:
        st.metric("Reducción % PIB global", f"{total_row['Reducción % PIB']*100:.2f}%")

    st.markdown("---")

    # Callout
    st.markdown("""
    <div class="callout">
    <h4>💡 Hallazgo clave</h4>
    <p>Una inversión global de <b>$36.5 mil millones</b> en sistemas de alerta temprana podría generar
    <b>$661 mil millones</b> en beneficios durante 10 años, con un <b>BCR de 7.2x</b> — consistente con
    el rango de 3:1 a 10:1 documentado por McKinsey y la ONU.</p>
    </div>
    """, unsafe_allow_html=True)

    # BCR por región
    st.subheader("BCR por Región")
    fig = px.bar(
        df_regiones.sort_values("BCR", ascending=True),
        x="BCR", y="Región", orientation="h",
        color="BCR",
        color_continuous_scale=["#A84B2F", "#FFC553", "#20808D"],
        text=df_regiones.sort_values("BCR", ascending=True)["BCR"].apply(lambda x: f"{x:.1f}x"),
    )
    fig.update_layout(height=400, showlegend=False, plot_bgcolor="#F7F6F2", paper_bgcolor="#F7F6F2")
    fig.update_traces(textposition="outside")
    fig.update_coloraxes(showscale=False)
    st.plotly_chart(fig, width="stretch")

    # Infografía
    st.markdown("---")
    st.subheader("🖼️ Infografía Resumen")
    st.image(str(IMG_INFOGRAFIA), caption="Infografía — Catástrofes vs Alerta Temprana", width="stretch")


# ──────────────────────────────────────────────
# PÁGINA: MODELO FINANCIERO
# ──────────────────────────────────────────────
elif pagina == "💰 Modelo Financiero":
    st.title("💰 Modelo Financiero por Región")
    st.markdown("### KPIs detallados — Escenario Base")
    st.markdown("---")

    # Tabla completa
    st.subheader("Tabla de KPIs por Región")
    df_display = df_resumen.copy()
    # Formatear columnas
    for col in ["Pérdidas Hist. ($B)", "Inversión EWS ($B)", "Pérdidas Evitadas 10a ($B)",
                "Ahorro FR ($B)", "Beneficio Total ($B)", "Coste Total EWS ($B)"]:
        df_display[col] = df_display[col].apply(lambda x: f"${x:.1f}B")
    df_display["ROI"] = df_display["ROI"].apply(lambda x: f"{x*100:.0f}%")
    df_display["BCR"] = df_display["BCR"].apply(lambda x: f"{x:.1f}x")
    df_display["Reducción % PIB"] = df_display["Reducción % PIB"].apply(lambda x: f"{x*100:.2f}%")
    st.dataframe(df_display, width="stretch", hide_index=True)

    st.markdown("---")

    # Gráficos interactivos
    col_a, col_b = st.columns(2)
    with col_a:
        st.subheader("Inversión EWS vs Beneficio Total")
        fig2 = go.Figure()
        fig2.add_trace(go.Bar(name="Inversión EWS", x=df_regiones["Región"],
                              y=df_regiones["Inversión EWS ($B)"], marker_color="#A84B2F"))
        fig2.add_trace(go.Bar(name="Beneficio Total", x=df_regiones["Región"],
                              y=df_regiones["Beneficio Total ($B)"], marker_color="#20808D"))
        fig2.update_layout(barmode="group", height=400, plot_bgcolor="#F7F6F2", paper_bgcolor="#F7F6F2")
        st.plotly_chart(fig2, width="stretch")

    with col_b:
        st.subheader("ROI por Región")
        fig3 = px.bar(
            df_regiones.sort_values("ROI", ascending=False),
            x="Región", y="ROI",
            color="ROI",
            color_continuous_scale=["#A84B2F", "#FFC553", "#20808D"],
            text=df_regiones.sort_values("ROI", ascending=False)["ROI"].apply(lambda x: f"{x*100:.0f}%"),
        )
        fig3.update_layout(height=400, showlegend=False, plot_bgcolor="#F7F6F2", paper_bgcolor="#F7F6F2")
        fig3.update_traces(textposition="outside")
        fig3.update_coloraxes(showscale=False)
        st.plotly_chart(fig3, width="stretch")

    st.markdown("---")
    st.subheader("Pérdidas Evitadas vs Coste EWS")
    fig4 = go.Figure()
    fig4.add_trace(go.Bar(name="Pérdidas Evitadas (10 años)", x=df_regiones["Región"],
                          y=df_regiones["Pérdidas Evitadas 10a ($B)"], marker_color="#20808D"))
    fig4.add_trace(go.Bar(name="Ahorro First Responders", x=df_regiones["Región"],
                          y=df_regiones["Ahorro FR ($B)"], marker_color="#FFC553"))
    fig4.add_trace(go.Bar(name="Coste Total EWS", x=df_regiones["Región"],
                          y=df_regiones["Coste Total EWS ($B)"], marker_color="#A84B2F"))
    fig4.update_layout(barmode="group", height=450, plot_bgcolor="#F7F6F2", paper_bgcolor="#F7F6F2")
    st.plotly_chart(fig4, width="stretch")


# ──────────────────────────────────────────────
# PÁGINA: PÉRDIDAS HISTÓRICAS
# ──────────────────────────────────────────────
elif pagina == "📉 Pérdidas Históricas":
    st.title("📉 Pérdidas Históricas por Catástrofes Naturales")
    st.markdown("### 2021-2025 — Datos por región ($B)")
    st.markdown("---")

    st.dataframe(df_historicos, width="stretch", hide_index=True)

    st.markdown("---")

    # Gráfico de líneas
    st.subheader("Evolución de Pérdidas por Región (2021-2025)")
    df_melt = df_historicos.melt(id_vars=["Región"], value_vars=["2021","2022","2023","2024","2025"],
                                  var_name="Año", value_name="Pérdidas ($B)")
    fig5 = px.line(df_melt, x="Año", y="Pérdidas ($B)", color="Región",
                   markers=True, color_discrete_sequence=CHART_PALETTE)
    fig5.update_layout(height=450, plot_bgcolor="#F7F6F2", paper_bgcolor="#F7F6F2")
    st.plotly_chart(fig5, width="stretch")

    # Stacked bar
    st.subheader("Composición de Pérdidas por Año")
    fig6 = px.bar(df_melt, x="Año", y="Pérdidas ($B)", color="Región",
                  color_discrete_sequence=CHART_PALETTE)
    fig6.update_layout(height=400, plot_bgcolor="#F7F6F2", paper_bgcolor="#F7F6F2")
    st.plotly_chart(fig6, width="stretch")

    # Promedio
    st.markdown("---")
    st.subheader("Promedio de Pérdidas por Región (2021-2025)")
    fig7 = px.bar(df_historicos.sort_values("Promedio", ascending=True),
                  x="Promedio", y="Región", orientation="h",
                  color="Promedio",
                  color_continuous_scale=["#BCE2E7", "#20808D", "#1B474D"],
                  text=df_historicos.sort_values("Promedio", ascending=True)["Promedio"].apply(lambda x: f"${x:.1f}B"))
    fig7.update_layout(height=400, showlegend=False, plot_bgcolor="#F7F6F2", paper_bgcolor="#F7F6F2")
    fig7.update_traces(textposition="outside")
    fig7.update_coloraxes(showscale=False)
    st.plotly_chart(fig7, width="stretch")


# ──────────────────────────────────────────────
# PÁGINA: PROYECCIONES
# ──────────────────────────────────────────────
elif pagina == "🔮 Proyecciones 10 años":
    st.title("🔮 Proyecciones a 10 Años por Región")
    st.markdown("### Flujo neto anual (2026-2035) — Escenario Base")
    st.markdown("---")

    # Selector de región
    region_sel = st.selectbox("Selecciona una región:", df_proyecciones["Región"].tolist())
    row_sel = df_proyecciones[df_proyecciones["Región"] == region_sel].iloc[0]

    years = [str(y) for y in range(2026, 2036)]
    valores = [row_sel[y] for y in years]

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Flujo neto total (10 años)", f"${sum(valores):.1f}B")
    with col2:
        st.metric("Flujo neto promedio anual", f"${sum(valores)/10:.1f}B")
    with col3:
        st.metric("Mejor año", f"${max(valores):.1f}B", str(2026 + valores.index(max(valores))))

    fig8 = go.Figure()
    fig8.add_trace(go.Bar(x=years, y=valores, marker_color="#20808D",
                          text=[f"${v:.1f}" for v in valores], textposition="outside"))
    fig8.update_layout(title=f"Flujo Neto Anual — {region_sel}",
                       xaxis_title="Año", yaxis_title="$B",
                       height=400, plot_bgcolor="#F7F6F2", paper_bgcolor="#F7F6F2")
    st.plotly_chart(fig8, width="stretch")

    # Heatmap de todas las regiones
    st.markdown("---")
    st.subheader("Mapa de Calor — Flujo Neto por Región y Año")
    df_heat = df_proyecciones.set_index("Región")
    fig9 = px.imshow(df_heat, color_continuous_scale=["#A84B2F", "#FFC553", "#20808D"],
                     labels=dict(x="Año", y="Región", color="$B"),
                     text_auto=".1f")
    fig9.update_layout(height=450, plot_bgcolor="#F7F6F2", paper_bgcolor="#F7F6F2")
    st.plotly_chart(fig9, width="stretch")

    # Tabla
    st.markdown("---")
    st.subheader("Tabla Detallada de Proyecciones")
    st.dataframe(df_proyecciones, width="stretch", hide_index=True)


# ──────────────────────────────────────────────
# PÁGINA: ESCENARIOS
# ──────────────────────────────────────────────
elif pagina == "📈 Comparativa de Escenarios":
    st.title("📈 Comparativa de Escenarios")
    st.markdown("### Conservador vs Base vs Optimista")
    st.markdown("---")

    # Tabla de escenarios
    esc_data = []
    for metric, vals in escenarios.items():
        esc_data.append({
            "Métrica": metric,
            "Conservador": vals["Conservador"],
            "Base": vals["Base"],
            "Optimista": vals["Optimista"],
            "Unidad": vals["Unidad"],
        })
    df_esc = pd.DataFrame(esc_data)
    st.dataframe(df_esc, width="stretch", hide_index=True)

    st.markdown("---")

    # Gráficos comparativos
    col_a, col_b = st.columns(2)

    with col_a:
        st.subheader("BCR por Escenario")
        bcr_row = escenarios["BCR (Beneficio/Coste)"]
        fig10 = go.Figure()
        fig10.add_trace(go.Bar(x=["Conservador", "Base", "Optimista"],
                               y=[bcr_row["Conservador"], bcr_row["Base"], bcr_row["Optimista"]],
                               marker_color=["#A84B2F", "#20808D", "#437A22"],
                               text=[f'{bcr_row["Conservador"]:.1f}x', f'{bcr_row["Base"]:.1f}x', f'{bcr_row["Optimista"]:.1f}x'],
                               textposition="outside"))
        fig10.update_layout(height=400, yaxis_title="BCR (x)",
                           plot_bgcolor="#F7F6F2", paper_bgcolor="#F7F6F2")
        # Línea de referencia 3:1
        fig10.add_hline(y=3, line_dash="dash", line_color="#964219",
                       annotation_text="Umbral McKinsey 3:1")
        st.plotly_chart(fig10, width="stretch")

    with col_b:
        st.subheader("ROI por Escenario")
        roi_row = escenarios["ROI proyectado"]
        fig11 = go.Figure()
        fig11.add_trace(go.Bar(x=["Conservador", "Base", "Optimista"],
                               y=[roi_row["Conservador"]*100, roi_row["Base"]*100, roi_row["Optimista"]*100],
                               marker_color=["#A84B2F", "#20808D", "#437A22"],
                               text=[f'{roi_row["Conservador"]*100:.0f}%', f'{roi_row["Base"]*100:.0f}%', f'{roi_row["Optimista"]*100:.0f}%'],
                               textposition="outside"))
        fig11.update_layout(height=400, yaxis_title="ROI (%)",
                           plot_bgcolor="#F7F6F2", paper_bgcolor="#F7F6F2")
        st.plotly_chart(fig11, width="stretch")

    st.markdown("---")
    st.subheader("Beneficio vs Coste por Escenario")
    ben_row = escenarios["Beneficio total 10 años ($B)"]
    cost_row = escenarios["Coste total EWS ($B)"]
    fig12 = go.Figure()
    fig12.add_trace(go.Bar(name="Beneficio Total", x=["Conservador", "Base", "Optimista"],
                           y=[ben_row["Conservador"], ben_row["Base"], ben_row["Optimista"]],
                           marker_color="#20808D"))
    fig12.add_trace(go.Bar(name="Coste Total EWS", x=["Conservador", "Base", "Optimista"],
                           y=[cost_row["Conservador"], cost_row["Base"], cost_row["Optimista"]],
                           marker_color="#A84B2F"))
    fig12.update_layout(barmode="group", height=400, yaxis_title="$B",
                        plot_bgcolor="#F7F6F2", paper_bgcolor="#F7F6F2")
    st.plotly_chart(fig12, width="stretch")

    st.markdown("---")
    st.subheader("Gráfico Comparativo Original (matplotlib)")
    st.image(str(IMG_CHART), caption="Gráfico de escenarios — 4 paneles", width="stretch")


# ──────────────────────────────────────────────
# PÁGINA: DOCUMENTACIÓN
# ──────────────────────────────────────────────
elif pagina == "📄 Documentación":
    st.title("📄 Documentación")
    st.markdown("---")

    st.subheader("PDF Resumen Ejecutivo")
    st.markdown("Documento de 4 páginas con KPIs, metodología, escenarios y conclusiones.")
    display_pdf(PDF_RESUMEN)

    st.markdown("---")
    st.subheader("Excel — Modelo Financiero Completo")
    st.markdown("6 hojas: Resumen, Supuestos, Datos Históricos, Escenarios, Proyecciones y Metodología.")
    file_download(XLSX, "📊 Descargar Excel completo")

    st.markdown("---")
    st.subheader("Infografía Resumen")
    st.image(str(IMG_INFOGRAFIA), caption="Infografía para inversores", width="stretch")


# ──────────────────────────────────────────────
# PÁGINA: ANÁLISIS DE SESGOS
# ──────────────────────────────────────────────
elif pagina == "⚠️ Análisis de Sesgos":
    st.title("⚠️ Análisis de Sesgos Socio-Político-Culturales")
    st.markdown("### Los 6 sesgos críticos del modelo que todo inversor debe conocer")
    st.markdown("---")

    st.markdown("""
    <div class="callout-warn">
    <h4>⚠️ Importante</h4>
    <p>Un modelo que reconoce sus sesgos es <b>más creíble</b> que uno que los oculta.
    El BCR conservador de 3.4x se mantiene incluso tras este análisis crítico,
    por encima del umbral de 3:1 de McKinsey.</p>
    </div>
    """, unsafe_allow_html=True)

    # KPIs de sesgos
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Desastres sin datos de daños", "41.5%", "EM-DAT")
    with col2:
        st.metric("Sin daños asegurados", "88.1%", "EM-DAT")
    with col3:
        st.metric("Mortalidad con EWS", "6-8x menor", "No en ROI")
    with col4:
        st.metric("PIB perdido en Jamaica", "40%", "1 huracán")

    st.markdown("---")

    # Infografía
    st.subheader("Infografía de Sesgos")
    st.image(str(IMG_SESGOS), caption="Análisis visual de los 6 sesgos críticos", width="stretch")

    st.markdown("---")

    # PDF completo
    st.subheader("PDF — Análisis Completo de Sesgos")
    st.markdown("7 páginas con análisis detallado de cada sesgo, fuentes académicas y recomendaciones de mitigación.")
    display_pdf(PDF_SESGOS)

    st.markdown("---")
    st.markdown("""
    <div class="callout">
    <h4>💡 Conclusión</h4>
    <p>Reconocer los sesgos <b>no debilita el caso de inversión</b>: incluso en el escenario conservador,
    el BCR de 3.4x supera el umbral mínimo de 3:1. La inclusión de estas salvedades fortalece la
    credibilidad del modelo y permite una toma de decisiones más informada y ética.</p>
    </div>
    """, unsafe_allow_html=True)
